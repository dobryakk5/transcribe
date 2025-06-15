# handlers_common.py
from aiogram.types import Message
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.types.input_file import BufferedInputFile
from aiogram import Bot, Dispatcher
from db_handler import update_dictionary, get_today_purchases, get_user_purchases, update_last_purchase_field, get_last_purchase, delete_last_purchase
from start_handlers import on_start
import pandas as pd
import textwrap
import matplotlib.pyplot as plt
from io import BytesIO
from datetime import datetime
import asyncpg
from asyncpg.exceptions import UniqueViolationError
import logging
import uuid
import redis
import os
from dotenv import load_dotenv

# шифровка и передача user_id
#from itsdangerous import URLSafeSerializer
#serializer = URLSafeSerializer(API_TOKEN, salt="uid-salt")

r = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)

# Инициализация бота
load_dotenv()
BOT_TOKEN = os.getenv("API_TOKEN")
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


# Инициализация логгера
logger = logging.getLogger(__name__)

async def show_parser_result(category: str, subcategory: str, price: str, message: Message):
    """Показывает результат работы парсера"""
    await message.answer(
        f"🤖 Парсер вернул:\n"
        f"• Категория: <b>{category}</b>\n"
        f"• Подкатегория: <b>{subcategory}</b>\n"
        f"• Цена: <b>{price}</b>",
        parse_mode="HTML"
    )

async def handle_correction(field: str, new_val: str, message: Message):
    # Telegram ID как строка для избежания проблем с большими числами
    user_id_str = str(message.from_user.id)
    
    field_names = {
        'category': 'Категория',
        'subcategory': 'Подкатегория',
        'price': 'Цена'
    }
    field_rus = field_names.get(field, field)
    
    await message.answer(f"✏️ Обновляю поле <b>{field_rus}</b> на «{new_val}»…", parse_mode="HTML")
    
    try:
        if field in ('category', 'subcategory'):
            # Для категорий и подкатегорий используем глобальное обновление справочника
            last_purchase = await get_last_purchase(user_id_str)
            if not last_purchase:
                await message.answer("⚠️ Нет предыдущей записи для обновления.")
                return
                
            if field == 'category':
                ok = await update_dictionary(
                    user_id_str,
                    'category',
                    old_name=last_purchase['category'],
                    new_name=new_val
                )
            else:
                ok = await update_dictionary(
                    user_id_str,
                    'subcategory',
                    category_name=last_purchase['category'],
                    old_name=last_purchase['subcategory'],
                    new_name=new_val
                )
        else:
            # Для цены используем точечное обновление
            ok = await update_last_purchase_field(user_id_str, field, new_val)
        
        if ok:
            await message.answer("✅ Поле обновлено.")
        else:
            await message.answer("⚠️ Нет предыдущей записи для обновления.")
            
    except ValueError as ve:
        await message.answer(f"❌ Ошибка значения: {ve}")
    except UniqueViolationError:
        await message.answer("❌ Такое значение уже существует в справочнике")
    except DataError as de:
        logger.error(f"DataError: {de}")
        await message.answer("❌ Ошибка обработки данных. Пожалуйста, попробуйте позже.")
    except Exception as e:
        logger.exception("Ошибка при обновлении")
        await message.answer(f"❌ Критическая ошибка: {str(e)}")

async def show_today_purchases(user_id: int, message: Message):
    rows = await get_today_purchases(user_id)

    if not rows:
        return await message.answer("Сегодня ещё нет записей.")

    # Формируем строки с подкатегорией и ценой (до 25 символов), цена с разделителями
    lines = [
        f"{r['subcategory'][:25]:<25} {int(r['price']):>10,}".replace(",", ".")
        for r in rows
    ]

    # Вычисляем общую сумму и добавляем итоговую строку
    total = sum(int(r['price']) for r in rows)
    lines.append("")
    lines.append(f"{'Итого сумма сегодня:':<25} {total:>10,}".replace(",", "."))

    await message.answer(f"<pre>{chr(10).join(lines)}</pre>", parse_mode="HTML")

async def export_purchases_to_excel(user_id: int, filename: str):
    """
    Экспортирует все покупки пользователя в Excel-файл.
    Если записей нет, создаёт файл с заголовками.
    """
    # Получаем записи, гарантируем список
    rows = await get_user_purchases(user_id) or []

    # Подготавливаем данные для DataFrame
    data = []
    for row in rows:
        price_str = f"{int(row['price']):,}".replace(",", ".")
        data.append({
            'Категория': row['category'],
            'Подкатегория': row['subcategory'],
            'Цена': price_str,
            'Дата': row['ts'].astimezone(tz=None).replace(tzinfo=None).strftime("%d.%m.%Y %H:%M")
        })

    # Если нет данных, создаём пустой DataFrame с колонками
    if not data:
        df = pd.DataFrame(columns=['Категория', 'Подкатегория', 'Цена', 'Дата'])
    else:
        df = pd.DataFrame(data)

    # Сохраняем в Excel
    df.to_excel(filename, index=False)

    # Форматирование Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(filename)
    ws = wb.active

    # Автоширина колонок
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Выравнивание цен и дат
    header = [cell.value for cell in ws[1]]
    if 'Цена' in header:
        price_col = header.index('Цена') + 1
        for cell in ws.iter_rows(min_row=2, min_col=price_col, max_col=price_col):
            cell[0].alignment = Alignment(horizontal='right')
    if 'Дата' in header:
        time_col = header.index('Дата') + 1
        for cell in ws.iter_rows(min_row=2, min_col=time_col, max_col=time_col):
            cell[0].alignment = Alignment(horizontal='right')

    wb.save(filename)

# Chart display function
async def show_pie_chart(user_id: int, message: Message):
    rows = await get_user_purchases(user_id)
    if not rows:
        await message.answer("Нет данных для построения графика.")
        return

    category_totals = {}
    for row in rows:
        cat = row["category"]
        price = float(row["price"])
        category_totals[cat] = category_totals.get(cat, 0) + price

    labels = list(category_totals.keys())
    sizes = list(category_totals.values())

    def autopct_format(values):
        def my_format(pct):
            total = sum(values)
            v = int(round(pct * total / 100.0))
            return f"{pct:.1f}%\n({v} ₽)"
        return my_format

    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=labels,
        autopct=autopct_format(sizes),
        startangle=90
    )
    ax.axis("equal")

    # Настройка размера и шрифта текста (опционально)
    plt.setp(autotexts, size=10, weight="light")

    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close(fig)

    await message.answer_photo(photo=BufferedInputFile(buffer.read(), filename="chart.png"))

async def show_bar_chart_by_day(user_id: int, message: Message):
    rows = await get_user_purchases(user_id)
    if not rows:
        await message.answer("Нет данных для построения графика.")
        return

    df = pd.DataFrame([dict(r) for r in rows])
    df["ts"] = pd.to_datetime(df["ts"])
    df["date"] = df["ts"].dt.date
    df["price"] = pd.to_numeric(df["price"], errors="coerce")

    grouped = df.groupby(["date", "category"])["price"].sum().unstack(fill_value=0)
    cumulative = grouped.cumsum()

    ax = cumulative.plot(kind="bar", stacked=True, figsize=(10, 6))
    ax.set_ylabel("Сумма")
    ax.set_xlabel("Дата")
    ax.set_xticklabels([d.strftime("%d.%m") for d in cumulative.index], rotation=0)
    ax.set_title("Кумулятивные оплаты по категориям")
    ax.legend(title="Категории", bbox_to_anchor=(1.05, 1), loc="upper left")

    fig = ax.get_figure()
    fig.tight_layout()
    buffer = BytesIO()
    fig.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close(fig)

    await message.answer_photo(photo=BufferedInputFile(buffer.read(), filename="bar_chart.png"))

async def show_daily_bar_chart(user_id: int, message: Message):
    rows = await get_user_purchases(user_id)
    if not rows:
        await message.answer("Нет данных для построения графика.")
        return

    df = pd.DataFrame([dict(r) for r in rows])
    df["ts"] = pd.to_datetime(df["ts"])
    df["date"] = df["ts"].dt.date
    df["price"] = pd.to_numeric(df["price"], errors="coerce")

    grouped = df.groupby(["date", "category"])["price"].sum().unstack(fill_value=0)

    ax = grouped.plot(kind="bar", stacked=True, figsize=(10, 6))
    ax.set_ylabel("Сумма")
    ax.set_xlabel("Дата")
    ax.set_xticklabels([d.strftime("%d.%m") for d in grouped.index], rotation=0)
    ax.set_title("Ежедневные траты по категориям")
    ax.legend(title="Категории", bbox_to_anchor=(1.05, 1), loc="upper left")

    fig = ax.get_figure()
    fig.tight_layout()
    buffer = BytesIO()
    fig.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close(fig)

    await message.answer_photo(photo=BufferedInputFile(buffer.read(), filename="daily_bar_chart.png"))


    rows = await get_user_purchases(user_id)

    data = []
    for row in rows:
        price_str = f"{int(row['price']):,}".replace(",", ".")
        data.append({
            'Категория': row['category'],
            'Подкатегория': row['subcategory'],
            'Цена': price_str,
            'Дата': row['ts'].astimezone(tz=None).replace(tzinfo=None).strftime("%d.%m.%Y %H:%M")
        })

    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)

    # Форматирование Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(filename)
    ws = wb.active

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    header = [cell.value for cell in ws[1]]
    price_col = header.index('Цена') + 1
    time_col = header.index('Дата') + 1

    for row in ws.iter_rows(min_row=2, min_col=price_col, max_col=price_col):
        row[0].alignment = Alignment(horizontal='right')

    for row in ws.iter_rows(min_row=2, min_col=time_col, max_col=time_col):
        row[0].alignment = Alignment(horizontal='right')

    wb.save(filename)


async def process_user_input(
    raw_text: str, 
    message: Message,
    handle_new_expense_func
):
    lower = raw_text.lower().strip()

    # Кнопка «Инструкция»
    if lower == "📘 инструкция":
        await message.answer(
        textwrap.dedent("""\
            💸 Добавить новую оплату: напиши «категория подкатегория цена».
            🧾 Загрузить позиции с чека: отправь фото QR-кода с чека
            🎙️ Загрузить простую транзакцию голосом: запиши голосовое.
            🛠️ Исправить поле в последней записи: введи "удали" или выбери новое значение
              – «Категория НовоеЗначение»
              – «Подкатегория НовоеЗначение»
              – «Цена НовоеЗначение»
            📋 Показать список сегодняшних оплат: меню «Список» 
            🔢 Выгрузить все оплаты в Excel: меню «Таблица» 
            💰 Добавить свои доходы: напиши "доход консультация 49505"
             - слово "доход" в начале говорит боту писать консультацию в доходы.

            Обновить бота 🔄: напиши /start 
        """)
        )
        return

    if lower == "📈 графики":
        await message.answer(
            "📊 Что показать?",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="🔘 Круг по категориям")],
                    [KeyboardButton(text="📊 Накопительно категория/день")],
                    [KeyboardButton(text="📊 Ежедневно категория/день")],
                    [KeyboardButton(text="🏠 Главное меню")]
                ],
                resize_keyboard=True
            )
        )
        return

    if lower == "💰 доходы":
        from db_handler import get_user_incomes_days
        rows = await get_user_incomes_days(message.from_user.id, 30)
        if not rows:
            return await message.answer("Доходов за последние 30 дней нет.")
        lines = [
            f"{r['source'][:25]:<25} {int(r['amount']):>10,}".replace(",", ".")
            for r in rows
        ]
        total = sum(int(r['amount']) for r in rows)
        lines.append("")
        lines.append(f"{'Итого за 30 дней:':<25} {total:>10,}".replace(",", "."))
        await message.answer(f"<pre>{chr(10).join(lines)}</pre>", parse_mode="HTML")
        return

    if lower == "🔘 круг по категориям":
        await show_pie_chart(message.from_user.id, message)
        return

    if lower == "📊 накопительно категория/день":
        await show_bar_chart_by_day(message.from_user.id, message)
        return

    if lower == "📊 ежедневно категория/день":
        await show_daily_bar_chart(message.from_user.id, message)
        return

    if lower == "🏠 главное меню":
        await on_start(message)
        return
    
    if lower in ("🚪 кабинет", "кабинет"):
        user_id = str(message.from_user.id)
        token = str(uuid.uuid4())

        r.setex(f"dash_token:{token}", 300, user_id)
        dash_url = f"https://ai5.space/auth?token={token}"

        # Создаем клавиатуру с кнопкой
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔓 Войти в кабинет", url=dash_url)]
        ])

        await message.answer(
            "🔒 Ваша ссылка для входа (действительна 5 минут):\n\n",
            reply_markup=keyboard,
            parse_mode="HTML",
            disable_web_page_preview=True
        )
        return

    if lower == "📄 список":
        await show_today_purchases(message.from_user.id, message)
        return
    
    if lower == "удали":
        success = await delete_last_purchase(message.from_user.id)
        if success:
            await message.answer("✅ Последняя запись успешно удалена.")
        else:
            await message.answer("⚠️ Нет записей для удаления.")
        return


    if lower == "🔢 таблица":
        import os
        filename = "Fin_a_bot.xlsx"
        await export_purchases_to_excel(message.from_user.id, filename)
        with open(filename, 'rb') as f:
            file_data = f.read()
        await message.answer_document(BufferedInputFile(file_data, filename))
        os.remove(filename)
        return

    correction_commands = {
        "категория": "category",
        "подкатегория": "subcategory",
        "цена": "price"
    }

    for prefix, field in correction_commands.items():
        if lower.startswith(prefix):
            parts = raw_text.split(maxsplit=1)
            if len(parts) < 2 or not parts[1].strip():
                return await message.answer(f"❌ Укажите значение после «{prefix.capitalize()}»")
            return await handle_correction(field, parts[1].strip(), message)

    await handle_new_expense_func(raw_text, message)

